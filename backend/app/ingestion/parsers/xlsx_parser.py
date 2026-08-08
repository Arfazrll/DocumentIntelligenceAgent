"""
DocIntel AI — XLSX Parser.

Parses Excel files preserving table structure.
"""

import logging
from pathlib import Path

from app.ingestion.parsers.docling_parser import ParsedElement

logger = logging.getLogger(__name__)

class XlsxParser:
    """Parser for Excel files using openpyxl."""

    def parse(self, file_path: Path) -> list[ParsedElement]:
        """
        Parse an Excel file, converting each sheet into structured elements.
        """
        logger.info(f"Parsing XLSX: {file_path.name}")
        from openpyxl import load_workbook

        elements: list[ParsedElement] = []

        try:
            wb = load_workbook(str(file_path), data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Convert sheet to markdown table
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    if any(c.strip() for c in cells):  # Skip empty rows
                        rows.append(cells)

                if not rows:
                    continue

                # Build markdown table
                md_lines = []

                # Header
                header = rows[0]
                md_lines.append("| " + " | ".join(header) + " |")
                md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")

                # Data rows
                for row in rows[1:]:
                    # Pad or truncate to match header length
                    padded = row + [""] * (len(header) - len(row))
                    md_lines.append("| " + " | ".join(padded[:len(header)]) + " |")

                table_content = "\n".join(md_lines)

                elements.append(
                    ParsedElement(
                        content=table_content,
                        element_type="table",
                        section_path=f"Sheet: {sheet_name}",
                        metadata={
                            "parser": "openpyxl",
                            "sheet_name": sheet_name,
                            "row_count": len(rows),
                            "col_count": len(rows[0]) if rows else 0,
                        },
                    )
                )

            wb.close()
            logger.info(f"XLSX parsed {len(elements)} sheets from {file_path.name}")
            return elements

        except Exception as e:
            logger.error(f"XLSX parsing failed for {file_path.name}: {e}")
            raise

# Singleton instance
xlsx_parser = XlsxParser()
